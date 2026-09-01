"""
Balak Progress Sync — Streamlit UI
=====================================
Enter your karyakar portal credentials, click Run, watch progress
stream into a live log panel, then download the resulting Excel file.

Credentials are never written to disk — they're only held in Streamlit's
session state for this browser session and passed straight into the
scraper function.

Run with:  streamlit run app.py
"""

import io
import math
import re
from datetime import datetime

import matplotlib
import matplotlib.pyplot as plt
import streamlit as st
from openpyxl import load_workbook

matplotlib.use("Agg")

from scraper import run_scraper, ScraperError

st.set_page_config(page_title="Balak Progress Sync", page_icon="📘", layout="wide")

# --- session state -----------------------------------------------------
for key, default in {
    "status": "idle",       # idle | running | done | error
    "output_path": None,
    "output_bytes": None,
    "output_filename": None,
    "error_message": None,
    "row_count": 0,
}.items():
    if key not in st.session_state:
        st.session_state[key] = default

THEME = {
    "orange": "#E67E22",
    "navy": "#0A2540",
    "maroon": "#7C1D1D",
    "dark_green": "#0E6B52",
    "text": "#1F2937",
    "bg": "#F8FAFC",
    "graph_bg": "#F7EFE1",
}


def _safe_float(value):
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _completed_videos(chapters_completed):
    completed = _safe_float(chapters_completed)
    if completed <= 3:
        return 0
    result = ((completed - 3) / 3) * 2
    if result.is_integer():
        return int(result)
    return int(math.ceil(result))


def _load_excel_data(file_path):
    workbook = load_workbook(file_path, read_only=True)
    return _sheet_to_rows(workbook.active)


def _load_excel_bytes(file_bytes):
    workbook = load_workbook(io.BytesIO(file_bytes), read_only=True)
    return _sheet_to_rows(workbook.active)


def _sheet_to_rows(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    data = []
    for row in rows[1:]:
        if not any(cell is not None and str(cell).strip() not in ("", "None") for cell in row):
            continue
        item = {}
        for idx, header in enumerate(headers):
            value = row[idx] if idx < len(row) else ""
            item[header] = value
        data.append(item)
    return data


def _fig_to_jpg(fig):
    buffer = io.BytesIO()
    fig.savefig(buffer, format="jpeg", dpi=300, bbox_inches="tight")
    buffer.seek(0)
    plt.close(fig)
    return buffer.getvalue()


def _build_stacked_chart(data, value_mode, title_text, subtitle_text, completed_color, pending_color, output_name):
    chart_rows = []
    for row in data:
        name = row.get("Balak Name", "") or "Unknown"
        completed = _safe_float(row.get("Chapters Completed", 0))
        expected = _safe_float(row.get("Chapters Expected", 0))

        if value_mode == "chapters":
            completed_value = completed
            pending_value = max(expected - completed, 0)
        else:
            completed_value = _completed_videos(completed)
            pending_value = max(_completed_videos(expected) - completed_value, 0)

        chart_rows.append({
            "name": name,
            "completed": completed_value,
            "pending": pending_value,
        })

    if not chart_rows:
        return None

    chart_rows = sorted(chart_rows, key=lambda item: (item["completed"], item["pending"]), reverse=True)
    names = [item["name"] for item in chart_rows]
    completed_values = [item["completed"] for item in chart_rows]
    pending_values = [item["pending"] for item in chart_rows]

    fig, ax = plt.subplots(figsize=(12, max(7, len(names) * 0.32)))
    fig.patch.set_facecolor(THEME["graph_bg"])
    ax.set_facecolor(THEME["graph_bg"])
    fig.subplots_adjust(left=0.28, right=0.98, top=0.82, bottom=0.08)
    fig.patch.set_alpha(1)
    ax.patch.set_alpha(1)

    y_pos = range(len(names))

    completed_bars = ax.barh(
        list(y_pos),
        completed_values,
        color=completed_color,
        height=0.8,
        label="Completed course",
    )
    pending_bars = ax.barh(
        list(y_pos),
        pending_values,
        left=completed_values,
        color=pending_color,
        height=0.8,
        label="Pending",
    )

    ax.set_yticks(list(y_pos))
    ax.set_yticklabels(names)
    ax.invert_yaxis()
    ax.xaxis.set_visible(False)
    ax.yaxis.grid(False)
    ax.xaxis.grid(False)
    ax.set_axisbelow(False)

    for spine in ax.spines.values():
        spine.set_visible(False)

    ax.tick_params(axis="y", labelsize=12, colors=THEME["navy"])
    ax.tick_params(axis="x", bottom=False, labelbottom=False)
    for label in ax.get_yticklabels():
        label.set_color(THEME["navy"])
        label.set_fontweight("bold")

    for bar in completed_bars:
        height = bar.get_height()
        width = bar.get_width()
        if width <= 0:
            continue
        x = bar.get_x() + width
        ax.text(
            x - 0.5,
            bar.get_y() + height / 2,
            f"{int(width)}",
            va="center",
            ha="right",
            color="white",
            fontsize=14,
            fontweight="bold",
        )

    for bar in pending_bars:
        height = bar.get_height()
        width = bar.get_width()
        if width <= 0:
            continue
        x = bar.get_x() + width
        ax.text(
            x - 0.5,
            bar.get_y() + height / 2,
            f"{int(width)}",
            va="center",
            ha="right",
            color="white",
            fontsize=14,
            fontweight="bold",
        )

    fig.text(0.105, 0.96, title_text, ha="left", va="top", fontsize=18, fontweight="bold", color=THEME["orange"])
    fig.text(0.105, 0.92, subtitle_text, ha="left", va="top", fontsize=12, fontweight="bold", color=THEME["navy"])
    ax.legend(loc="upper center", bbox_to_anchor=(0.5, 1.08), ncol=2, frameon=False, fontsize=9, handlelength=1.8)

    for patch in ax.patches:
        patch.set_linewidth(1.2)
        patch.set_edgecolor("white")

    for spine in ax.spines.values():
        spine.set_visible(False)

    ax.margins(y=0.02)
    fig.subplots_adjust(left=0.30, right=0.98, top=0.82, bottom=0.08)

    return fig, output_name

# --- main layout ----------------------------------------------------------
left_col, right_col = st.columns([1.0, 1.5], gap="large")

with left_col:
    st.title("Balak Progress Sync")
    st.caption("Pulls SDVC chapter-completion data for every balak on your list into one spreadsheet.")
    st.caption("Credentials")
    with st.form("run_form"):
        username = st.text_input("Portal username")
        password = st.text_input("Portal password", type="password")
        submitted = st.form_submit_button("Run Scraper", use_container_width=True)

    st.caption("Your credentials are used only for this run and are never saved.")

with right_col:
    st.subheader("Process log")
    progress_bar = right_col.progress(0, text="Scrape progress")
    log_box = right_col.container(height=430, border=True)
    log_placeholder = log_box.empty()
    if st.session_state.status == "idle":
        log_placeholder.info("Run the scraper to see live progress here.")
    else:
        log_placeholder.caption("Running scraper...")

# --- run the scraper (blocking, with a live-updating log panel) ------------
if submitted:
    if not username or not password:
        st.error("Please enter both a username and password.")
    else:
        st.session_state.status = "running"
        st.session_state.output_path = None
        st.session_state.output_bytes = None
        st.session_state.output_filename = None
        st.session_state.error_message = None

        log_lines = []

        def update_progress_from_log(message: str):
            total_match = re.search(r"Found\s+(\d+)\s+balak report link", message)
            if total_match:
                progress_bar.progress(0.05, text="Preparing to scrape...")
                return

            item_match = re.search(r"\((\d+)\/((?:\d+))\)", message)
            if item_match:
                current, total = map(int, item_match.groups())
                percent = min(max(current / total, 0.05), 0.98)
                progress_bar.progress(percent, text=f"Scraping balak {current}/{total}")
                return

            if "Writing " in message and "row(s)" in message:
                progress_bar.progress(0.95, text="Writing Excel file...")
                return

            if message.rstrip(".") == "Done":
                progress_bar.progress(1.0, text="Completed")

        def log(message: str):
            log_lines.append(message)
            update_progress_from_log(message)

            log_placeholder.empty()
            with log_placeholder.container():
                for line in log_lines:
                    lower = line.lower()
                    if line.startswith("ERROR"):
                        st.markdown(f":red[{line}]")
                    elif lower.startswith("warning"):
                        st.markdown(f":orange[{line}]")
                    elif lower.rstrip(".") == "done":
                        st.markdown(f":green[**{line}**]")
                    else:
                        st.text(line)

        now = datetime.now()
        timestamp_label = now.strftime("%d_%b_%y_%H_%M_%S")
        report_filename = f"SDVC_Report_{timestamp_label}.xlsx"
        excel_buffer = io.BytesIO()

        try:
            run_scraper(username, password, excel_buffer, log, headless=True)
            st.session_state.status = "done"
            st.session_state.output_path = None
            st.session_state.output_filename = report_filename
            st.session_state.output_bytes = excel_buffer.getvalue()
        except ScraperError as e:
            st.session_state.status = "error"
            st.session_state.error_message = str(e)
            log(f"ERROR: {e}")
        except Exception as e:  # noqa: BLE001 - surface anything unexpected in the UI too
            st.session_state.status = "error"
            st.session_state.error_message = f"Unexpected error: {e}"
            log(f"ERROR: Unexpected error: {e}")

# --- result / download -----------------------------------------------------
if st.session_state.status == "done" and st.session_state.output_bytes:
    st.success("Done — your spreadsheet is ready.")
    st.download_button(
        "Download Excel",
        data=st.session_state.output_bytes,
        file_name=st.session_state.output_filename or f"SDVC_Report_{datetime.now().strftime('%d_%b_%y_%H_%M_%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    chart_data = _load_excel_bytes(st.session_state.output_bytes)
    if chart_data:
        chart_tabs = st.tabs([
            "Balak Name vs Chapters Completed",
            "Balak Name vs Completed Videos",
        ])

        with chart_tabs[0]:
            fig_data = _build_stacked_chart(
                chart_data,
                value_mode="chapters",
                title_text="Satsang Diksha Visharad",
                subtitle_text="Year 1",
                completed_color=THEME["navy"],
                pending_color=THEME["maroon"],
                output_name="chapters",
            )
            if fig_data:
                fig, _ = fig_data
                st.pyplot(fig)
                st.download_button(
                    "Download JPG",
                    data=_fig_to_jpg(fig),
                    file_name=f"SDVC_GRAPHNAME_{datetime.now().strftime('%d_%b_%y_%H_%M_%S')}.jpg",
                    mime="image/jpeg",
                    use_container_width=True,
                    key="download_jpg_chapters",
                )

        with chart_tabs[1]:
            fig_data = _build_stacked_chart(
                chart_data,
                value_mode="videos",
                title_text="Satsang Diksha Visharad",
                subtitle_text="Year 1",
                completed_color=THEME["dark_green"],
                pending_color=THEME["maroon"],
                output_name="videos",
            )
            if fig_data:
                fig, _ = fig_data
                st.pyplot(fig)
                st.download_button(
                    "Download JPG",
                    data=_fig_to_jpg(fig),
                    file_name=f"SDVC_GRAPHNAME_{datetime.now().strftime('%d_%b_%y_%H_%M_%S')}.jpg",
                    mime="image/jpeg",
                    use_container_width=True,
                    key="download_jpg_videos",
                )

elif st.session_state.status == "error" and st.session_state.error_message:
    st.error(st.session_state.error_message)
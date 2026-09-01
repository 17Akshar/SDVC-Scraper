"""
Balak Progress Scraper — core logic (UI-agnostic)
====================================================
Same scraping logic as the CLI version, refactored so the web app can
call it in a background thread and stream progress to the browser
instead of printing to a terminal.

Nothing here writes credentials to disk — username/password are passed
in as plain function arguments and only ever live in memory for the
duration of one run.
"""

import math
import os
import shutil
import time
from dataclasses import dataclass

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


LOGIN_URL = "https://balmandal.in.baps.org/Secure/Login.aspx?returnurl=%2f"
WAIT_TIMEOUT = 20


class LOC:
    USERNAME_INPUT = (By.XPATH, '//*[@id="ctl01_mainContent_LoginCtrl_UserName"]')
    PASSWORD_INPUT = (By.XPATH, '//*[@id="ctl01_mainContent_LoginCtrl_Password"]')
    LOGIN_BUTTON = (By.XPATH, '//*[@id="ctl01_mainContent_LoginCtrl_Login"]')

    POST_LOGIN_IMAGE_BUTTON = (By.XPATH, '//*[@id="ctl01_mainContent_ctl01_ImageButton1"]')

    NIRDESHAK_TILE = (
        By.XPATH,
        "/html/body/div/div[2]/section/div/div[2]/div/div/div[2]/div/div[3]/a",
    )

    CARDS = (By.CSS_SELECTOR, "div.sdvc-card")
    CARD_REPORT_LINK = (By.CSS_SELECTOR, 'a[title="sdvc"]')
    TOTAL_COUNT_BADGE = (By.CSS_SELECTOR, "#total-count")
    MANDAL_SELECT = (By.CSS_SELECTOR, "select#mandal")
    NEXT_PAGE_LINK = (
        By.CSS_SELECTOR,
        "a.page-link[rel='next'], li.next:not(.disabled) a, a[aria-label='Next']",
    )

    BALAK_NAME = (By.CSS_SELECTOR, "div.sdvc-hero h4")
    COURSE_TILES = (By.CSS_SELECTOR, ".course-tile")
    COURSE_TITLE = (By.CSS_SELECTOR, "h5")
    STAT_NUMS = (By.CSS_SELECTOR, ".stat-num")
    STATUS_NOTE = (By.CSS_SELECTOR, ".status-note")


@dataclass
class ProgressRow:
    name: str
    report_url: str
    course: str = ""
    completed: str = ""
    expected: str = ""
    status: str = ""


class ScraperError(Exception):
    """Raised for any failure we want surfaced cleanly to the UI."""


def _wait_for(driver, locator, timeout=WAIT_TIMEOUT, clickable=False):
    condition = EC.element_to_be_clickable if clickable else EC.presence_of_element_located
    return WebDriverWait(driver, timeout).until(condition(locator))


def _get_chrome_binary():
    candidates = [
        os.environ.get("CHROME_BIN"),
        os.environ.get("GOOGLE_CHROME_BIN"),
        "/usr/bin/chromium",
        "/usr/bin/chromium-browser",
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
        "/usr/local/bin/chromium",
        "/usr/local/bin/google-chrome",
        "google-chrome",
        "google-chrome-stable",
        "chromium",
        "chromium-browser",
        "chrome",
    ]
    for candidate in candidates:
        if not candidate:
            continue
        resolved = shutil.which(candidate)
        if resolved:
            return resolved
        if os.path.exists(candidate):
            return candidate
    return None


def _get_chromedriver_path():
    candidates = [
        os.environ.get("CHROMEDRIVER_PATH"),
        "/usr/bin/chromedriver",
        "/usr/local/bin/chromedriver",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "chromedriver"),
        shutil.which("chromedriver"),
    ]
    for candidate in candidates:
        if not candidate:
            continue
        if os.path.exists(candidate) and os.access(candidate, os.X_OK):
            return candidate
    return None


def _get_driver(headless: bool):
    chrome_binary = _get_chrome_binary()
    driver_path = _get_chromedriver_path()

    if not chrome_binary:
        raise RuntimeError(
            "Chrome/Chromium browser is not installed in this deployment environment. "
            "Install Chromium via Streamlit packages.txt or use a custom Docker image with Chrome."
        )
    if not driver_path:
        raise RuntimeError(
            "Chromedriver is not available in this deployment environment. "
            "Add chromium-driver in packages.txt or upload a Linux x86_64 chromedriver file to the repo."
        )

    options = Options()
    options.binary_location = chrome_binary

    if headless:
        options.add_argument("--headless=new")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-gpu")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--disable-extensions")

    service = Service(driver_path)
    driver = webdriver.Chrome(service=service, options=options)
    driver.implicitly_wait(2)
    return driver


def _login(driver, username, password, log):
    log("Logging in...")
    driver.get(LOGIN_URL)
    _wait_for(driver, LOC.USERNAME_INPUT).send_keys(username)
    driver.find_element(*LOC.PASSWORD_INPUT).send_keys(password)
    _wait_for(driver, LOC.LOGIN_BUTTON, clickable=True).click()
    time.sleep(2)
    log("Logged in.")


def _click_post_login_image(driver, log):
    log("Opening SDVC portal...")
    btn = _wait_for(driver, LOC.POST_LOGIN_IMAGE_BUTTON, clickable=True)
    btn.click()
    WebDriverWait(driver, WAIT_TIMEOUT).until(EC.url_contains("sdmukhpath.karyakar.in"))
    log(f"Redirected to {driver.current_url}")


def _click_nirdeshak_tile(driver, log):
    log("Loading balak list...")
    tile = _wait_for(driver, LOC.NIRDESHAK_TILE, clickable=True)
    tile.click()
    time.sleep(2)
    _wait_for(driver, LOC.CARDS)
    log("Balak list loaded.")


def _ensure_all_mandals_selected(driver, log):
    try:
        select_el = driver.find_element(*LOC.MANDAL_SELECT)
    except NoSuchElementException:
        return
    if select_el.get_attribute("value") == "All":
        return
    driver.execute_script(
        """
        const el = arguments[0];
        el.value = 'All';
        el.dispatchEvent(new Event('change', { bubbles: true }));
        """,
        select_el,
    )
    time.sleep(1.5)
    _wait_for(driver, LOC.CARDS)
    log("Mandal filter set to 'All'.")


def _collect_report_urls(driver, log):
    _ensure_all_mandals_selected(driver, log)
    _wait_for(driver, LOC.CARDS)

    urls = []
    page_num = 1
    while True:
        cards = driver.find_elements(*LOC.CARDS)
        for card in cards:
            try:
                link = card.find_element(*LOC.CARD_REPORT_LINK)
                href = link.get_attribute("href")
                if href and href not in urls:
                    urls.append(href)
            except NoSuchElementException:
                continue

        next_links = driver.find_elements(*LOC.NEXT_PAGE_LINK)
        if not next_links:
            break

        log(f"Page {page_num} done ({len(cards)} card(s)) — moving to next page...")
        next_links[0].click()
        time.sleep(1.5)
        _wait_for(driver, LOC.CARDS)
        page_num += 1

    try:
        expected_total = int(driver.find_element(*LOC.TOTAL_COUNT_BADGE).text.strip())
        if expected_total != len(urls):
            log(f"Warning: page shows SDVC Count {expected_total}, "
                f"but found {len(urls)} report links total.")
    except (NoSuchElementException, ValueError):
        pass

    return urls


def _scrape_report_page(driver, url):
    driver.get(url)
    name = _wait_for(driver, LOC.BALAK_NAME).text.strip()
    tiles = driver.find_elements(*LOC.COURSE_TILES)

    if not tiles:
        return [ProgressRow(name=name, report_url=url, status="No progress data found")]

    rows = []
    for tile in tiles:
        try:
            course_name = tile.find_element(*LOC.COURSE_TITLE).text.strip()
        except NoSuchElementException:
            course_name = ""
        stat_nums = tile.find_elements(*LOC.STAT_NUMS)
        completed = stat_nums[0].text.strip() if len(stat_nums) > 0 else ""
        expected = stat_nums[1].text.strip() if len(stat_nums) > 1 else ""
        try:
            status = tile.find_element(*LOC.STATUS_NOTE).text.strip()
        except NoSuchElementException:
            status = ""
        rows.append(ProgressRow(name, url, course_name, completed, expected, status))
    return rows


def _completed_videos(chapters_completed):
    try:
        completed = float(chapters_completed)
    except (TypeError, ValueError):
        return 0

    if completed <= 3:
        return 0

    result = ((completed - 3) / 3) * 2
    return int(math.ceil(result)) if not result.is_integer() else int(result)


def save_to_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balak Progress"

    headers = [
        "Balak Name",
        "Course",
        "Chapters Completed",
        "Completed Videos",
        "Chapters Expected",
        "Status",
        "Report URL",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header).font = Font(bold=True)

    for row_idx, r in enumerate(rows, start=2):
        completed_videos = _completed_videos(r.completed)
        ws.cell(row=row_idx, column=1, value=r.name)
        ws.cell(row=row_idx, column=2, value=r.course)
        ws.cell(row=row_idx, column=3, value=r.completed)
        ws.cell(row=row_idx, column=4, value=completed_videos)
        ws.cell(row=row_idx, column=5, value=r.expected)
        ws.cell(row=row_idx, column=6, value=r.status)
        ws.cell(row=row_idx, column=7, value=r.report_url)

    for col, width in enumerate([26, 32, 18, 18, 18, 30, 45], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)


def run_scraper(username: str, password: str, output_path: str, log, headless: bool = True):
    """
    Runs the full scrape end-to-end and writes an Excel file to output_path.
    `log` is a callable — call log(message) to report progress; the web
    app wires this up to stream lines to the browser in real time.
    Raises ScraperError with a UI-friendly message on failure.
    """
    if not username or not password:
        raise ScraperError("Username and password are required.")

    try:
        driver = _get_driver(headless=headless)
    except Exception as e:
        raise ScraperError(
            "Browser startup failed. Chrome/Chromium is not installed in this deployment environment. "
            "Use a custom Docker image with Chrome or a VM/server that supports Selenium browser automation."
        ) from e

    try:
        _login(driver, username, password, log)

        # Basic check: did login actually succeed, or are we still on the login page?
        if "Login.aspx" in driver.current_url:
            raise ScraperError("Login failed — please check your username and password.")

        _click_post_login_image(driver, log)
        _click_nirdeshak_tile(driver, log)

        urls = _collect_report_urls(driver, log)
        log(f"Found {len(urls)} balak report link(s). Scraping...")

        rows = []
        for i, url in enumerate(urls, start=1):
            try:
                page_rows = _scrape_report_page(driver, url)
                for r in page_rows:
                    label = f"{r.completed}/{r.expected}" if r.completed else r.status
                    log(f"({i}/{len(urls)}) {r.name} [{r.course or '-'}]: {label}")
                rows.extend(page_rows)
            except (TimeoutException, NoSuchElementException) as e:
                log(f"Failed on {url}: {e}")
                rows.append(ProgressRow(name="UNKNOWN", report_url=url, status=f"ERROR: {e}"))

        log(f"Writing {len(rows)} row(s) to Excel...")
        save_to_excel(rows, output_path)
        log("Done.")
        return output_path

    except TimeoutException as e:
        raise ScraperError(f"Timed out waiting for a page element: {e}")
    finally:
        driver.quit()